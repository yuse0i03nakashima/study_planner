import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database import get_connection
from planner import build_plan_data


def _write_excel_sheet(ws, subjects, rows, unassigned,
                       student_name, start_date, end_date):
    """全教科を1シートに出力するダークテーマExcel"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date as date_cls

    # ── カラーパレット ──
    BG_MAIN     = "0C0D11"
    BG_SURFACE  = "13151E"
    BG_SURFACE2 = "1B1E2B"
    BG_SURFACE3 = "222536"
    BG_DAY_A    = "13151E"   # 日付ブロックA（濃）
    BG_DAY_B    = "1B1E2B"   # 日付ブロックB（やや薄）
    C_TEXT      = "DDE1EC"
    C_MUTED     = "9AA3B8"
    C_DIM       = "555D7A"
    C_BLUE      = "5B8FF9"
    C_GREEN     = "3ECF8E"
    C_AMBER     = "F5A623"
    C_ROSE      = "F06292"
    C_RED       = "EF4444"
    C_BORDER    = "252838"
    C_BORDER_L  = "2F3347"

    CAT_COLORS = {
        "予習": C_BLUE, "復習": C_GREEN,
        "定着": C_AMBER, "再定着": C_ROSE,
    "New":       C_BLUE,
    "Recall":    C_GREEN,
    "Drill":     C_AMBER,
    "Reinforce": C_ROSE,
    }
    MASTERY_COLORS = {1: C_ROSE, 2: C_AMBER, 3: C_GREEN}
    DOW_JA = ["月", "火", "水", "木", "金", "土", "日"]

    def fill(c): return PatternFill("solid", fgColor=c)
    def font(c, bold=False, size=10, mono=False):
        name = "Consolas" if mono else "Noto Serif JP"
        return Font(color=c, bold=bold, size=size, name=name)
    def aln(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def bdr(c=C_BORDER):
        s = Side(style="thin", color=c)
        return Border(left=s, right=s, top=s, bottom=s)

    ws.sheet_view.showGridLines = False

    # ── 列定義 ──
    # Date / Day / Subject / Category / Textbook / Problem / Mastery / Min / Instruction
    COLS = [
        ("Date",        10),
        ("Day",          4),
        ("Subject",     10),
        ("Category",     9),
        ("Textbook",    20),
        ("Problem",     34),
        ("Mastery",      8),
        ("Min",          5),
        ("Instruction", 44),
    ]
    N_COLS = len(COLS)

    # ── 全セルにデフォルト背景色を設定 ──
    # I列より右（J〜Z列）にもダークテーマを適用
    for r in range(1, 501):
        for c in range(1, 40):  # A〜AN列まで
            ws.cell(row=r, column=c).fill = fill(BG_MAIN)

    # ── 列幅 ──
    for i, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── タイトル行 ──
    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    c1 = ws["A1"]
    c1.value = f"[SP]  {student_name}   {start_date} → {end_date}"
    c1.fill = fill("0F1119")   # タイトル専用（最も濃い）
    c1.font = Font(color=C_BLUE, bold=True, size=12, name="Consolas")
    c1.alignment = aln("left", "center")
    ws.row_dimensions[1].height = 30

    # ── ヘッダ行 ──
    BG_HEADER_ROW = "222536"   # ヘッダ専用色（BG_SURFACE3相当）
    for i, (h, _) in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=i)
        cell.value = h.upper()
        cell.fill = fill(BG_HEADER_ROW)
        cell.font = Font(color=C_MUTED, bold=False, size=8, name="Consolas")
        cell.alignment = aln("center", "center")
        cell.border = bdr(C_BORDER_L)
    ws.row_dimensions[2].height = 20

    ws.freeze_panes = "A3"

    # ── データ行 ──
    cur = 3
    day_toggle = False  # 日付ブロックの交互色制御

    for day_row in rows:
        date_str_raw = day_row.get("date_str", "")
        date_label   = day_row.get("date", "")
        subjects_dict = day_row.get("subjects", {})

        # この日の全問題を収集（教科順）
        day_items = []
        for subj in subjects:
            for item in subjects_dict.get(subj, []):
                day_items.append((subj, item))

        if not day_items:
            continue

        # 日付ブロックの背景色を交互に
        day_bg = BG_DAY_A if day_toggle else BG_DAY_B
        day_toggle = not day_toggle

        try:
            d_obj = date_cls.fromisoformat(date_str_raw)
            dow = DOW_JA[d_obj.weekday()]
            date_disp = f"{d_obj.month}/{d_obj.day}"
            is_weekend = d_obj.weekday() >= 5
        except Exception:
            dow = ""
            date_disp = date_label
            is_weekend = False

        # 曜日色
        dow_color = C_ROSE if is_weekend else C_MUTED

        for row_i, (subj, item) in enumerate(day_items):
            cat = item.get("category", "")
            cat_color = CAT_COLORS.get(cat, C_DIM)
            mastery_int = item.get("mastery_int", 1)
            mastery_color = MASTERY_COLORS.get(mastery_int, C_TEXT)

            # HP進捗バーをProblem列に組み込む
            _prob_text = str(item.get("problem_number", ""))
            _stot = int(item.get("session_total", 1) or 1)
            _sidx = int(item.get("session_index", 1) or 1)
            if _stot > 1:
                _pa2  = float(item.get("progress_after", 1.0) or 1.0)
                _fil  = int(_pa2 * 8)
                _bar2 = "█" * _fil + "░" * (8 - _fil)
                _pct2 = int(_pa2 * 100)
                _prob_text = (_prob_text + "  [" + _bar2 + "]"
                              + str(_pct2) + "% ("
                              + str(_sidx) + "/" + str(_stot) + ")")

            vals = [
                date_disp if row_i == 0 else "",
                dow       if row_i == 0 else "",
                subj,
                cat,
                item.get("textbook", ""),
                _prob_text,
                item.get("mastery", "★"),
                item.get("estimated_minutes", ""),
                item.get("instruction", ""),
            ]

            # 同一日の最終行かどうか（下ボーダーを強調）
            is_last_in_day = (row_i == len(day_items) - 1)
            bot_side = Side(style="medium", color=C_BORDER_L)                        if is_last_in_day else Side(style="thin", color=C_BORDER)
            top_side = Side(style="thin", color=C_BORDER)
            side_side = Side(style="thin", color=C_BORDER)

            for col_i, val in enumerate(vals, 1):
                cell = ws.cell(row=cur, column=col_i)
                cell.value = val
                cell.fill = fill(day_bg)
                cell.border = Border(
                    left=side_side, right=side_side,
                    top=top_side, bottom=bot_side
                )

                # 列別スタイル
                if col_i == 1:   # Date
                    cell.font = font(C_MUTED, size=9, mono=True)
                    cell.alignment = aln("center", "center")
                elif col_i == 2: # Day
                    cell.font = Font(color=dow_color, size=9, name="Consolas")
                    cell.alignment = aln("center", "center")
                elif col_i == 3: # Subject
                    cell.font = font(C_MUTED, size=9, mono=True)
                    cell.alignment = aln("center", "center")
                elif col_i == 4: # Category
                    cell.font = Font(color=cat_color, bold=True,
                                     size=9, name="Consolas")
                    cell.alignment = aln("center", "center")
                elif col_i == 5: # Textbook
                    cell.font = font(C_MUTED, size=9)
                    cell.alignment = aln("left", "center")
                elif col_i == 6: # Problem
                    cell.font = font(C_TEXT, size=10)
                    cell.alignment = aln("left", "center", wrap=True)
                elif col_i == 7: # Mastery
                    cell.font = Font(color=mastery_color, size=10,
                                     name="Noto Serif JP")
                    cell.alignment = aln("center", "center")
                elif col_i == 8: # Min
                    cell.font = font(C_DIM, size=9, mono=True)
                    cell.alignment = aln("center", "center")
                elif col_i == 9: # Instruction
                    cell.font = font(C_MUTED, size=9)
                    cell.alignment = aln("left", "center", wrap=True)

            # 行の高さを内容に応じて動的設定（_prob_textはvals生成時に計算済み）
            instruction_text = str(item.get("instruction", ""))
            # 問題番号：列幅34で1行≒20文字
            prob_lines = max(1, -(-len(_prob_text) // 20))
            # 学習指示：列幅44で1行≒25文字
            instr_lines = max(1, -(-len(instruction_text) // 25))
            row_h = max(prob_lines, instr_lines) * 15 + 4
            row_h = max(20, min(row_h, 200))  # 20〜200ptの範囲
            ws.row_dimensions[cur].height = row_h
            cur += 1

    # ── 未割当ブロック ──
    all_unassigned = []
    for subj in subjects:
        for item in unassigned.get(subj, []):
            all_unassigned.append((subj, item))

    if all_unassigned:
        # 区切り
        ws.merge_cells(f"A{cur}:{get_column_letter(N_COLS)}{cur}")
        sep = ws.cell(row=cur, column=1)
        sep.value = "── UNASSIGNED ──"
        sep.fill = fill("1A0A0A")
        sep.font = Font(color=C_RED, size=8, name="Consolas")
        sep.alignment = aln("left", "center")
        ws.row_dimensions[cur].height = 14
        cur += 1

        for subj, item in all_unassigned:
            cat = item.get("category", "")
            cat_color = CAT_COLORS.get(cat, C_DIM)
            vals = ["—", "—", subj, cat,
                    item.get("textbook", ""),
                    item.get("problem_number", ""),
                    "—",
                    item.get("estimated_minutes", ""),
                    item.get("instruction", "")]
            for col_i, val in enumerate(vals, 1):
                cell = ws.cell(row=cur, column=col_i)
                cell.value = val
                cell.fill = fill("1A0A0A")
                cell.font = font(C_DIM, size=9)
                cell.border = bdr("3A1010")
                cell.alignment = aln("center" if col_i in (1,2,7,8) else "left",
                                     "center")
            ws.row_dimensions[cur].height = 18
            cur += 1

    # オートフィルター
    ws.auto_filter.ref = f"A2:{get_column_letter(N_COLS)}{cur - 1}"


def export_excel(student_id, start_date_str, end_date_str, subject_filter=None, section_ids=None):
    """計画表をExcelファイルに出力して保存パスを返す"""
    import openpyxl
    import os
    from datetime import datetime

    plan_data = build_plan_data(
        student_id, start_date_str, end_date_str, subject_filter)
    if not plan_data:
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    student_name = plan_data["student_name"]
    subjects     = plan_data["subjects"]
    rows         = plan_data["rows"]
    unassigned   = plan_data.get("unassigned", {})

    # 全教科を1シートに出力
    ws = wb.create_sheet(title="Plan")
    _write_excel_sheet(ws, subjects, rows, unassigned,
                       student_name, start_date_str, end_date_str)

    # 保存
    archive_dir = os.path.join(os.path.dirname(__file__), "plan_archives")
    os.makedirs(archive_dir, exist_ok=True)
    base_name = f"{student_name}_{start_date_str}_{end_date_str}"
    filename  = f"{base_name}.xlsx"
    filepath  = os.path.join(archive_dir, filename)
    # 同名ファイルが開かれている場合はタイムスタンプを付加
    if os.path.exists(filepath):
        try:
            import tempfile
            with open(filepath, "a"):
                pass
        except PermissionError:
            from datetime import datetime as _dt
            filename = f"{base_name}_{_dt.now().strftime('%H%M%S')}.xlsx"
            filepath = os.path.join(archive_dir, filename)
    wb.save(filepath)

    # historyテーブルに記録
    try:
        conn = get_connection()
        c = conn.cursor()
        c.execute("""
            INSERT INTO plan_history (student_id, filename, created_at)
            VALUES (?, ?, ?)
        """, (student_id, filename,
              datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit()
        conn.close()
    except Exception:
        pass

    return filepath

