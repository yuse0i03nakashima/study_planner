import os

base = r"C:\Users\ynaka\study_planner"

# ─── 1. update_masteryにscheduled_dateを追加 ────────────
mcp_path = os.path.join(base, "mcp_server.py")
with open(mcp_path, "r", encoding="utf-8") as f:
    mcp = f.read()

old_mastery_tool = '''        Tool(
            name="update_mastery",
            description="生徒の問題に対する習熟度を手動で修正する。誤った記録の修正や手動調整に使用。mastery: 1=要定着、2=定着中、3=定着済",
            inputSchema={
                "type": "object",
                "properties": {
                    "student_id": {"type": "string", "description": "生徒ID"},
                    "problem_id": {"type": "integer", "description": "問題ID"},
                    "mastery": {"type": "integer", "description": "新しい習熟度（1〜3）"},
                    "update_assignment": {"type": "boolean", "description": "出題予定のカテゴリも自動更新するか（省略時true）"}
                },
                "required": ["student_id", "problem_id", "mastery"]
            }
        ),'''

new_mastery_tool = '''        Tool(
            name="update_mastery",
            description="生徒の問題に対する習熟度を手動で修正する。誤った記録の修正や手動調整に使用。mastery: 1=要定着、2=定着中、3=定着済。scheduled_dateを指定すると次回出題日も上書きできる。",
            inputSchema={
                "type": "object",
                "properties": {
                    "student_id": {"type": "string", "description": "生徒ID"},
                    "problem_id": {"type": "integer", "description": "問題ID"},
                    "mastery": {"type": "integer", "description": "新しい習熟度（1〜3）"},
                    "scheduled_date": {"type": "string", "description": "次回出題日を手動指定（YYYY-MM-DD形式・省略時は自動計算）"},
                    "update_assignment": {"type": "boolean", "description": "出題予定のカテゴリも自動更新するか（省略時true）"}
                },
                "required": ["student_id", "problem_id", "mastery"]
            }
        ),'''

mcp = mcp.replace(old_mastery_tool, new_mastery_tool)

old_mastery_impl = '''        if update_assignment:
            c.execute("SELECT review_value FROM problems WHERE problem_id=?",
                      (problem_id,))
            p_row = c.fetchone()
            if p_row:
                review_value = p_row["review_value"]
                next_date = get_next_date(review_value, new_mastery, today)
                category = {1: "復習", 2: "定着"}.get(new_mastery, "再定着")
                c.execute("DELETE FROM assignments WHERE student_id=? AND problem_id=?",
                          (student_id, problem_id))
                c.execute("""
                    INSERT INTO assignments (student_id, problem_id, scheduled_date, category)
                    VALUES (?, ?, ?, ?)
                """, (student_id, problem_id, next_date.isoformat(), category))'''

new_mastery_impl = '''        if update_assignment:
            c.execute("SELECT review_value FROM problems WHERE problem_id=?",
                      (problem_id,))
            p_row = c.fetchone()
            if p_row:
                review_value = p_row["review_value"]
                scheduled_date = arguments.get("scheduled_date", "")
                if scheduled_date:
                    next_date_str = scheduled_date
                else:
                    next_date = get_next_date(review_value, new_mastery, today)
                    next_date_str = next_date.isoformat()
                category = {1: "復習", 2: "定着"}.get(new_mastery, "再定着")
                c.execute("DELETE FROM assignments WHERE student_id=? AND problem_id=?",
                          (student_id, problem_id))
                c.execute("""
                    INSERT INTO assignments (student_id, problem_id, scheduled_date, category)
                    VALUES (?, ?, ?, ?)
                """, (student_id, problem_id, next_date_str, category))'''

mcp = mcp.replace(old_mastery_impl, new_mastery_impl)

with open(mcp_path, "w", encoding="utf-8") as f:
    f.write(mcp)
print("✅ mcp_server.py を更新しました（update_mastery拡張）")

# ─── 2. excel_export.pyの優先度・分散ロジックを改修 ─────
excel_code = '''import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database import (get_connection, get_plan_v2, get_schedule,
                      get_class_dates_in_range, get_next_class_date)
from datetime import date, timedelta
from collections import defaultdict

DOW_JA = ["月", "火", "水", "木", "金", "土", "日"]

CATEGORY_COLORS = {
    "予習":   "DDEEFF",
    "復習":   "DDFFDD",
    "定着":   "FFFADD",
    "再定着": "FFE5DD",
    "未割当": "F0F0F0",
}

MASTERY_MULTIPLIER = {1: 1.3, 2: 1.0, 3: 0.5}

# カテゴリの優先グループ
CATEGORY_GROUP = {"復習": 0, "定着": 0, "再定着": 0, "予習": 1}


def get_adjusted_minutes(item):
    mastery_level = item.get("mastery_int", 1)
    mastery_level = max(1, min(3, mastery_level))
    multiplier = MASTERY_MULTIPLIER.get(mastery_level, 1.0)
    base_minutes = item["estimated_minutes"]
    return max(5, round(base_minutes * multiplier / 5) * 5)


def priority_score(item):
    """
    優先度スコアを計算する（高いほど先に配置）
    カテゴリグループ：復習・定着・再定着（0）> 予習（1）
    同グループ内：習熟度低い順、復習価値高い順、重要度高い順、難易度高い順
    """
    group = CATEGORY_GROUP.get(item["category"], 1)
    mastery = item.get("mastery_int", 1)
    review_value = item.get("review_value", 3)
    importance = item.get("importance", 3)
    difficulty = item.get("difficulty", 3)

    # スコアが高いほど優先（マイナスにして降順ソートに対応）
    return (
        group,           # 0=復習系優先、1=予習
        mastery,         # 1<2<3（低い習熟度を優先）
        -review_value,   # 高い復習価値を優先
        -importance,     # 高い重要度を優先
        -difficulty,     # 高い難易度を優先（難しい問題を先に）
    )


def assign_days_v2(plan, schedule, student_id, start_date_str, end_date_str):
    """
    優先度×テキスト分散を考慮した割り振りロジック

    1. 優先度スコアで全問題をソート
       （復習・定着・再定着 > 予習、同内では習熟度低>復習価値高>重要度高>難易度高）
    2. 復習は授業直後の最初の日に配置
    3. 予習は授業直前に配置
    4. 定着・再定着は残り時間にテキスト分散しながら配置
    5. テキスト分散：同じ日に同じテキストが集中しないよう調整
    """
    remaining = {d: m for d, m in schedule.items() if m > 0}
    dates_sorted = sorted(remaining.keys())

    if not dates_sorted:
        for item in plan:
            item["assigned_date"] = None
        return plan, []

    # 教科ごとの授業日を取得
    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT DISTINCT p.subject FROM problems p
        JOIN assignments a ON p.problem_id = a.problem_id
        WHERE a.student_id=?
    """, (student_id,))
    subjects = [r["subject"] for r in c.fetchall()]
    conn.close()

    subject_class_dates = {}
    for subject in subjects:
        subject_class_dates[subject] = get_class_dates_in_range(
            student_id, subject, start_date_str, end_date_str)

    assigned = []
    unassigned = []

    # 日付ごとのテキスト使用状況を追跡
    date_textbook_count = defaultdict(lambda: defaultdict(int))

    def try_assign(item, search_dates):
        """指定された日付候補から割り振り先を探す（テキスト分散考慮）"""
        minutes = get_adjusted_minutes(item)
        textbook = item["textbook"]

        # まずテキストが少ない日から試みる
        sorted_dates = sorted(
            [d for d in search_dates if remaining.get(d, 0) >= minutes],
            key=lambda d: date_textbook_count[d][textbook]
        )
        for d in sorted_dates:
            remaining[d] -= minutes
            item["assigned_date"] = d
            date_textbook_count[d][textbook] += 1
            assigned.append(item)
            return True
        return False

    # 全問題を優先度スコアでソート
    plan_sorted = sorted(plan, key=priority_score)

    # 復習・定着・再定着と予習を分離
    review_teichaku = [p for p in plan_sorted
                       if p["category"] in ("復習", "定着", "再定着")]
    yosyu = [p for p in plan_sorted if p["category"] == "予習"]

    # ─── Step1：復習を授業直後の最初の日に配置 ───
    for item in [p for p in review_teichaku if p["category"] == "復習"]:
        subject = item["subject"]
        class_dates = subject_class_dates.get(subject, [])

        # 授業直後の日から探す
        search_from = dates_sorted[0]
        past_classes = [d for d in class_dates if d <= start_date_str]
        if past_classes:
            search_from = past_classes[-1]

        search_dates = [d for d in dates_sorted if d >= search_from]
        if not try_assign(item, search_dates):
            item["assigned_date"] = None
            unassigned.append(item)

    # ─── Step2：定着・再定着を全体にテキスト分散して配置 ───
    for item in [p for p in review_teichaku
                 if p["category"] in ("定着", "再定着")]:
        if not try_assign(item, dates_sorted):
            item["assigned_date"] = None
            unassigned.append(item)

    # ─── Step3：予習を授業直前に配置 ───
    for item in yosyu:
        subject = item["subject"]
        class_dates = subject_class_dates.get(subject, [])

        future_classes = sorted([d for d in class_dates if d > start_date_str])
        if future_classes:
            next_class = future_classes[0]
            search_dates = list(reversed(
                [d for d in dates_sorted if d <= next_class]))
        else:
            search_dates = list(reversed(dates_sorted))

        if not try_assign(item, search_dates):
            item["assigned_date"] = None
            unassigned.append(item)

    return assigned, unassigned


def make_cell_text(item):
    line1 = f"【{item[\'category\']}】{item[\'textbook\']} {item[\'problem_number\']}"
    if item["instruction"]:
        return line1 + "\\n" + item["instruction"]
    return line1


def export_excel(target_date_str, output_path, student_id=None, start_date=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    conn = get_connection()
    c = conn.cursor()
    if student_id:
        c.execute("SELECT student_id, name, subjects FROM students "
                  "WHERE student_id=?", (student_id,))
    else:
        c.execute("SELECT student_id, name, subjects FROM students "
                  "ORDER BY student_id")
    students = c.fetchall()
    conn.close()

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for student in students:
        sid = student["student_id"]
        name = student["name"]
        subjects = [s.strip() for s in student["subjects"].split(",")]

        ws = wb.create_sheet(title=name)

        start_date_str = start_date if start_date else date.today().isoformat()
        schedule = get_schedule(sid, start_date_str, target_date_str)
        plan = get_plan_v2(sid, start_date_str, target_date_str)
        assigned, unassigned = assign_days_v2(
            plan, schedule, sid, start_date_str, target_date_str)

        dates_with_time = sorted([d for d, m in schedule.items() if m > 0])

        headers = ["実施日"] + subjects + ["未割当"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row_idx, d in enumerate(dates_with_time, start=2):
            d_obj = date.fromisoformat(d)
            date_label = (f"{d_obj.month}/{d_obj.day}"
                          f"（{DOW_JA[d_obj.weekday()]}）")

            date_cell = ws.cell(row=row_idx, column=1, value=date_label)
            date_cell.font = Font(bold=True)
            date_cell.fill = PatternFill("solid", fgColor="E8EEF8")
            date_cell.alignment = Alignment(
                horizontal="center", vertical="top")
            date_cell.border = border

            day_items = [item for item in assigned
                         if item["assigned_date"] == d]

            for col_idx, subject in enumerate(subjects, start=2):
                subject_items = [item for item in day_items
                                 if item["subject"] == subject]
                # 表示順：復習→定着→再定着→予習
                order = {"復習": 0, "定着": 1, "再定着": 2, "予習": 3}
                subject_items.sort(key=lambda x: order.get(x["category"], 9))

                cell_lines = [make_cell_text(item) for item in subject_items]
                cell_text = "\\n".join(cell_lines)

                color = "FFFFFF"
                if subject_items:
                    color = CATEGORY_COLORS.get(
                        subject_items[0]["category"], "FFFFFF")

                cell = ws.cell(row=row_idx, column=col_idx, value=cell_text)
                if subject_items:
                    cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True)
                cell.border = border

            unassigned_cell = ws.cell(
                row=row_idx, column=len(subjects) + 2, value="")
            unassigned_cell.border = border

        if unassigned:
            unassigned_row = len(dates_with_time) + 2
            label_cell = ws.cell(row=unassigned_row, column=1, value="未割当")
            label_cell.font = Font(bold=True)
            label_cell.fill = PatternFill("solid", fgColor="F0F0F0")
            label_cell.alignment = Alignment(
                horizontal="center", vertical="top")
            label_cell.border = border

            for col_idx, subject in enumerate(subjects, start=2):
                subject_items = [item for item in unassigned
                                 if item["subject"] == subject]
                cell_lines = [make_cell_text(item) for item in subject_items]
                cell_text = "\\n".join(cell_lines)
                cell = ws.cell(row=unassigned_row, column=col_idx,
                               value=cell_text)
                cell.fill = PatternFill("solid", fgColor="F0F0F0")
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True)
                cell.border = border

        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, len(subjects) + 3):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)].width = 35
        for row_idx in range(2, len(dates_with_time) + 3):
            ws.row_dimensions[row_idx].height = 80

    wb.save(output_path)
    return output_path
'''

excel_path = os.path.join(base, "excel_export.py")
with open(excel_path, "w", encoding="utf-8") as f:
    f.write(excel_code)
print("✅ excel_export.py を更新しました（優先度・テキスト分散ロジック改修）")

print("✅ 全て完了")