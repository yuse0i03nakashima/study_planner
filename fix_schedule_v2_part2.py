import os

base = r"C:\Users\ynaka\study_planner"
templates = os.path.join(base, "templates")

# ─── excel_export.py の改修 ────────────────────────────
excel_code = '''import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database import (get_connection, get_plan_v2, get_schedule,
                      get_class_dates_in_range, get_next_class_date)
from datetime import date, timedelta

DOW_JA = ["月", "火", "水", "木", "金", "土", "日"]

CATEGORY_COLORS = {
    "予習":   "DDEEFF",
    "復習":   "DDFFDD",
    "定着":   "FFFADD",
    "再定着": "FFE5DD",
    "未割当": "F0F0F0",
}

MASTERY_MULTIPLIER = {1: 1.3, 2: 1.0, 3: 0.5}

def get_adjusted_minutes(item):
    mastery_level = item.get("mastery_int", 1)
    mastery_level = max(1, min(3, mastery_level))
    multiplier = MASTERY_MULTIPLIER.get(mastery_level, 1.0)
    base_minutes = item["estimated_minutes"]
    return max(5, round(base_minutes * multiplier / 5) * 5)


def assign_days_v2(plan, schedule, student_id, start_date_str, end_date_str):
    """
    新しい割り振りロジック：
    1. 復習：授業直後の最初の利用可能日に配置
    2. 予習：次回授業日の直前（前日優先）に配置
    3. 定着・再定着：残り時間に全体にちりばめて配置
    """
    remaining = {d: m for d, m in schedule.items() if m > 0}
    dates_sorted = sorted(remaining.keys())

    if not dates_sorted:
        for item in plan:
            item["assigned_date"] = None
        return plan, []

    # 教科ごとの授業日を取得
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT DISTINCT subject FROM problems p "
              "JOIN assignments a ON p.problem_id = a.problem_id "
              "WHERE a.student_id=?", (student_id,))
    subjects = [r["subject"] for r in c.fetchall()]
    conn.close()

    subject_class_dates = {}
    for subject in subjects:
        subject_class_dates[subject] = get_class_dates_in_range(
            student_id, subject, start_date_str, end_date_str)

    assigned = []
    unassigned = []

    # ─── Step1：復習を授業直後の最初の日に配置 ───
    review_items = [p for p in plan if p["category"] == "復習"]
    for item in review_items:
        subject = item["subject"]
        class_dates = subject_class_dates.get(subject, [])
        minutes = get_adjusted_minutes(item)
        placed = False

        # 授業直後の日から探す
        search_from = dates_sorted[0]
        if class_dates:
            past_classes = [d for d in class_dates if d <= start_date_str]
            if past_classes:
                search_from = past_classes[-1]

        for d in dates_sorted:
            if d >= search_from and remaining.get(d, 0) >= minutes:
                remaining[d] -= minutes
                item["assigned_date"] = d
                assigned.append(item)
                placed = True
                break

        if not placed:
            item["assigned_date"] = None
            unassigned.append(item)

    # ─── Step2：予習を授業直前の日に配置 ───
    yosyu_items = [p for p in plan if p["category"] == "予習"]
    for item in yosyu_items:
        subject = item["subject"]
        class_dates = subject_class_dates.get(subject, [])
        minutes = get_adjusted_minutes(item)
        placed = False

        # 次回授業日の前日から逆順に探す
        future_classes = sorted([d for d in class_dates if d > start_date_str])
        search_dates = list(reversed(dates_sorted))
        if future_classes:
            next_class = future_classes[0]
            search_dates = list(reversed(
                [d for d in dates_sorted if d <= next_class]))

        for d in search_dates:
            if remaining.get(d, 0) >= minutes:
                remaining[d] -= minutes
                item["assigned_date"] = d
                assigned.append(item)
                placed = True
                break

        if not placed:
            item["assigned_date"] = None
            unassigned.append(item)

    # ─── Step3：定着・再定着を残り時間にちりばめる ───
    teichaku_items = [p for p in plan
                      if p["category"] in ("定着", "再定着")]
    # 復習価値・重要度の高い順に並べる
    teichaku_items.sort(key=lambda x: (
        -x["review_value"], -x["importance"]))

    # 各日に均等にちりばめるため、日付を循環させながら配置
    date_cycle = list(dates_sorted)
    cycle_idx = 0

    for item in teichaku_items:
        minutes = get_adjusted_minutes(item)
        placed = False
        attempts = 0

        while attempts < len(date_cycle):
            d = date_cycle[cycle_idx % len(date_cycle)]
            cycle_idx += 1
            attempts += 1
            if remaining.get(d, 0) >= minutes:
                remaining[d] -= minutes
                item["assigned_date"] = d
                assigned.append(item)
                placed = True
                break

        if not placed:
            item["assigned_date"] = None
            unassigned.append(item)

    return assigned, unassigned


def make_cell_text(item):
    line1 = f"【{item[\'category\']}】{item[\'textbook\']} {item[\'problem_number\']}"
    if item["instruction"]:
        return line1 + "\\n" + item["instruction"]
    return line1


def export_excel(target_date_str, output_path, student_id=None, start_date=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    conn = get_connection()
    c = conn.cursor()
    if student_id:
        c.execute("SELECT student_id, name, subjects FROM students "
                  "WHERE student_id=?", (student_id,))
    else:
        c.execute("SELECT student_id, name, subjects FROM students "
                  "ORDER BY student_id")
    students = c.fetchall()
    conn.close()

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for student in students:
        sid = student["student_id"]
        name = student["name"]
        subjects = [s.strip() for s in student["subjects"].split(",")]

        ws = wb.create_sheet(title=name)

        if start_date:
            start_date_str = start_date
        else:
            start_date_str = date.today().isoformat()

        schedule = get_schedule(sid, start_date_str, target_date_str)
        plan = get_plan_v2(sid, start_date_str, target_date_str)
        assigned, unassigned = assign_days_v2(
            plan, schedule, sid, start_date_str, target_date_str)

        dates_with_time = sorted([d for d, m in schedule.items() if m > 0])

        headers = ["実施日"] + subjects + ["未割当"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row_idx, d in enumerate(dates_with_time, start=2):
            d_obj = date.fromisoformat(d)
            date_label = (f"{d_obj.month}/{d_obj.day}"
                          f"（{DOW_JA[d_obj.weekday()]}）")

            date_cell = ws.cell(row=row_idx, column=1, value=date_label)
            date_cell.font = Font(bold=True)
            date_cell.fill = PatternFill("solid", fgColor="E8EEF8")
            date_cell.alignment = Alignment(
                horizontal="center", vertical="top")
            date_cell.border = border

            day_items = [item for item in assigned
                         if item["assigned_date"] == d]

            for col_idx, subject in enumerate(subjects, start=2):
                subject_items = [item for item in day_items
                                 if item["subject"] == subject]
                # 表示順：復習→定着→再定着→予習
                order = {"復習": 0, "定着": 1, "再定着": 2, "予習": 3}
                subject_items.sort(key=lambda x: order.get(x["category"], 9))

                cell_lines = [make_cell_text(item)
                              for item in subject_items]
                cell_text = "\\n".join(cell_lines)

                if subject_items:
                    color = CATEGORY_COLORS.get(
                        subject_items[0]["category"], "FFFFFF")
                else:
                    color = "FFFFFF"

                cell = ws.cell(row=row_idx, column=col_idx,
                               value=cell_text)
                if subject_items:
                    cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True)
                cell.border = border

            unassigned_cell = ws.cell(
                row=row_idx, column=len(subjects) + 2, value="")
            unassigned_cell.border = border

        if unassigned:
            unassigned_row = len(dates_with_time) + 2
            label_cell = ws.cell(row=unassigned_row, column=1,
                                 value="未割当")
            label_cell.font = Font(bold=True)
            label_cell.fill = PatternFill("solid", fgColor="F0F0F0")
            label_cell.alignment = Alignment(
                horizontal="center", vertical="top")
            label_cell.border = border

            for col_idx, subject in enumerate(subjects, start=2):
                subject_items = [item for item in unassigned
                                 if item["subject"] == subject]
                cell_lines = [make_cell_text(item)
                              for item in subject_items]
                cell_text = "\\n".join(cell_lines)
                cell = ws.cell(row=unassigned_row, column=col_idx,
                               value=cell_text)
                cell.fill = PatternFill("solid", fgColor="F0F0F0")
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True)
                cell.border = border

        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, len(subjects) + 3):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)].width = 35
        for row_idx in range(2, len(dates_with_time) + 3):
            ws.row_dimensions[row_idx].height = 80

    wb.save(output_path)
    return output_path
'''

with open(os.path.join(base, "excel_export.py"), "w", encoding="utf-8") as f:
    f.write(excel_code)
print("✅ excel_export.py を書き直しました")

# ─── app.pyに授業スケジュール管理ルートを追加 ──────────
app_path = os.path.join(base, "app.py")
with open(app_path, "r", encoding="utf-8") as f:
    app_content = f.read()

class_schedule_route = '''
@app.route("/class_schedule", methods=["GET", "POST"])
def class_schedule():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM students ORDER BY student_id")
    students = c.fetchall()

    if request.method == "POST":
        action = request.form.get("action")
        student_id = request.form["student_id"]
        subject = request.form["subject"]

        if action == "save_base":
            # 既存を削除して再登録
            c.execute("DELETE FROM class_schedule_base "
                      "WHERE student_id=? AND subject=?",
                      (student_id, subject))
            dows = request.form.getlist("dows")
            for dow in dows:
                c.execute("""
                    INSERT OR IGNORE INTO class_schedule_base
                    (student_id, subject, dow) VALUES (?, ?, ?)
                """, (student_id, subject, int(dow)))
            conn.commit()

        elif action == "save_override":
            next_date = request.form["next_class_date"]
            c.execute("""
                INSERT INTO class_schedule_override
                (student_id, subject, next_class_date)
                VALUES (?, ?, ?)
                ON CONFLICT(student_id, subject)
                DO UPDATE SET next_class_date=?
            """, (student_id, subject, next_date, next_date))
            conn.commit()

        elif action == "delete_override":
            c.execute("DELETE FROM class_schedule_override "
                      "WHERE student_id=? AND subject=?",
                      (student_id, subject))
            conn.commit()

        conn.close()
        return redirect(f"/class_schedule?student_id={student_id}")

    selected_student = request.args.get(
        "student_id", students[0]["student_id"] if students else None)

    # 現在の設定を取得
    base_schedule = {}
    override_schedule = {}
    student_subjects = []

    if selected_student:
        c.execute("SELECT subjects FROM students WHERE student_id=?",
                  (selected_student,))
        row = c.fetchone()
        if row:
            student_subjects = [s.strip()
                                 for s in row["subjects"].split(",")]

        c.execute("""
            SELECT subject, dow FROM class_schedule_base
            WHERE student_id=? ORDER BY subject, dow
        """, (selected_student,))
        for r in c.fetchall():
            base_schedule.setdefault(r["subject"], []).append(r["dow"])

        c.execute("""
            SELECT subject, next_class_date FROM class_schedule_override
            WHERE student_id=?
        """, (selected_student,))
        override_schedule = {r["subject"]: r["next_class_date"]
                             for r in c.fetchall()}

    conn.close()
    return render_template("class_schedule.html",
                           students=students,
                           selected_student=selected_student,
                           student_subjects=student_subjects,
                           base_schedule=base_schedule,
                           override_schedule=override_schedule)

'''

if "/class_schedule" not in app_content:
    idx = app_content.rfind("if __name__")
    app_content = app_content[:idx] + class_schedule_route + app_content[idx:]
    with open(app_path, "w", encoding="utf-8") as f:
        f.write(app_content)
    print("✅ app.py に授業スケジュールルートを追加しました")
else:
    print("ℹ️ 授業スケジュールルートはすでに存在します")

print("✅ Part2 完了")