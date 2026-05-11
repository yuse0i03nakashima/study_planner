import os
import sqlite3

base = r"C:\Users\ynaka\study_planner"
templates = os.path.join(base, "templates")

# ─── 1. studentsテーブルにplan_modeカラムを追加 ──────────
DB_PATH = os.path.join(base, "study_planner.db")
conn = sqlite3.connect(DB_PATH)
c = conn.cursor()
try:
    c.execute("ALTER TABLE students ADD COLUMN plan_mode TEXT DEFAULT 'all'")
    print("✅ plan_mode カラムを追加しました")
except sqlite3.OperationalError:
    print("ℹ️ plan_mode カラムはすでに存在します")

# 各生徒のデフォルト設定
c.execute("UPDATE students SET plan_mode='all' WHERE student_id='S001'")
c.execute("UPDATE students SET plan_mode='per_subject' WHERE student_id='S003'")
c.execute("UPDATE students SET plan_mode='per_subject' WHERE student_id='S004'")
conn.commit()
conn.close()
print("✅ 生徒のplan_modeを設定しました")

# ─── 2. excel_export.pyを更新 ──────────────────────────
excel_code = '''import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database import (get_connection, get_plan_v2, get_schedule,
                      get_class_dates_in_range)
from datetime import date, timedelta
from collections import defaultdict

DOW_JA = ["月", "火", "水", "木", "金", "土", "日"]

CATEGORY_COLORS = {
    "予習":   "DDEEFF",
    "復習":   "DDFFDD",
    "定着":   "FFFADD",
    "再定着": "FFE5DD",
    "未割当": "F0F0F0",
}

MASTERY_MULTIPLIER = {1: 1.3, 2: 1.0, 3: 0.5}
CATEGORY_GROUP = {"復習": 0, "定着": 0, "再定着": 0, "予習": 1}


def get_adjusted_minutes(item):
    mastery_level = item.get("mastery_int", 1)
    mastery_level = max(1, min(3, mastery_level))
    multiplier = MASTERY_MULTIPLIER.get(mastery_level, 1.0)
    return max(5, round(item["estimated_minutes"] * multiplier / 5) * 5)


def priority_score(item):
    group = CATEGORY_GROUP.get(item["category"], 1)
    mastery = item.get("mastery_int", 1)
    review_value = item.get("review_value", 3)
    importance = item.get("importance", 3)
    difficulty = item.get("difficulty", 3)
    return (group, mastery, -review_value, -importance, difficulty)


def assign_days_v2(plan, schedule, student_id, start_date_str, end_date_str):
    remaining = {d: m for d, m in schedule.items() if m > 0}
    dates_sorted = sorted(remaining.keys())
    if not dates_sorted:
        for item in plan:
            item["assigned_date"] = None
        return plan, []

    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT DISTINCT p.subject FROM problems p
        JOIN assignments a ON p.problem_id = a.problem_id
        WHERE a.student_id=?
    """, (student_id,))
    subjects = [r["subject"] for r in c.fetchall()]
    conn.close()

    subject_class_dates = {
        s: get_class_dates_in_range(student_id, s, start_date_str, end_date_str)
        for s in subjects
    }

    assigned = []
    unassigned = []
    date_textbook_count = defaultdict(lambda: defaultdict(int))

    def try_assign(item, search_dates):
        minutes = get_adjusted_minutes(item)
        textbook = item["textbook"]
        sorted_dates = sorted(
            [d for d in search_dates if remaining.get(d, 0) >= minutes],
            key=lambda d: date_textbook_count[d][textbook]
        )
        for d in sorted_dates:
            remaining[d] -= minutes
            item["assigned_date"] = d
            date_textbook_count[d][textbook] += 1
            assigned.append(item)
            return True
        return False

    plan_sorted = sorted(plan, key=priority_score)
    review_teichaku = [p for p in plan_sorted
                       if p["category"] in ("復習", "定着", "再定着")]
    yosyu = [p for p in plan_sorted if p["category"] == "予習"]

    for item in [p for p in review_teichaku if p["category"] == "復習"]:
        class_dates = subject_class_dates.get(item["subject"], [])
        search_from = dates_sorted[0]
        past_classes = [d for d in class_dates if d <= start_date_str]
        if past_classes:
            search_from = past_classes[-1]
        search_dates = [d for d in dates_sorted if d >= search_from]
        if not try_assign(item, search_dates):
            item["assigned_date"] = None
            unassigned.append(item)

    for item in [p for p in review_teichaku
                 if p["category"] in ("定着", "再定着")]:
        if not try_assign(item, dates_sorted):
            item["assigned_date"] = None
            unassigned.append(item)

    for item in yosyu:
        class_dates = subject_class_dates.get(item["subject"], [])
        future_classes = sorted([d for d in class_dates if d > start_date_str])
        if future_classes:
            next_class = future_classes[0]
            search_dates = list(reversed(
                [d for d in dates_sorted if d <= next_class]))
        else:
            search_dates = list(reversed(dates_sorted))
        if not try_assign(item, search_dates):
            item["assigned_date"] = None
            unassigned.append(item)

    return assigned, unassigned


def make_cell_text(item):
    line1 = f"【{item[\'category\']}】{item[\'textbook\']} {item[\'problem_number\']}"
    if item["instruction"]:
        return line1 + "\\n" + item["instruction"]
    return line1


def write_sheet_all(ws, sid, subjects, start_date_str, target_date_str,
                    border, thin):
    """全教科一括モード：1シートに全教科の列"""
    schedule = get_schedule(sid, start_date_str, target_date_str)
    plan = get_plan_v2(sid, start_date_str, target_date_str)
    assigned, unassigned = assign_days_v2(
        plan, schedule, sid, start_date_str, target_date_str)
    dates_with_time = sorted([d for d, m in schedule.items() if m > 0])

    headers = ["実施日"] + subjects + ["未割当"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    _write_rows(ws, dates_with_time, subjects, assigned, unassigned, border)


def write_sheet_per_subject(wb, sid, name, subjects, start_date_str,
                             target_date_str, border):
    """教科ごと個別モード：教科ごとにシートを作成"""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT dow, available_minutes FROM schedule_base "
              "WHERE student_id=?", (sid,))
    base_dow = {r["dow"]: r["available_minutes"] for r in c.fetchall()}
    c.execute("SELECT date, available_minutes FROM schedule_override "
              "WHERE student_id=? AND date BETWEEN ? AND ?",
              (sid, start_date_str, target_date_str))
    overrides = {r["date"]: r["available_minutes"] for r in c.fetchall()}
    conn.close()

    for subject in subjects:
        ws = wb.create_sheet(title=f"{name}_{subject}")

        # 教科の授業がある曜日のみスケジュールを生成
        class_dates = get_class_dates_in_range(
            sid, subject, start_date_str, target_date_str)

        # 教科の授業日を含む期間のスケジュール
        schedule = {}
        current = date.fromisoformat(start_date_str)
        end = date.fromisoformat(target_date_str)
        while current <= end:
            date_str = current.isoformat()
            schedule[date_str] = overrides.get(
                date_str, base_dow.get(current.weekday(), 0))
            current += timedelta(days=1)

        # この教科の問題のみを取得
        all_plan = get_plan_v2(sid, start_date_str, target_date_str)
        subject_plan = [p for p in all_plan if p["subject"] == subject]

        if not subject_plan:
            ws.cell(row=1, column=1, value=f"{subject}の出題予定はありません")
            continue

        assigned, unassigned = assign_days_v2(
            subject_plan, schedule, sid, start_date_str, target_date_str)
        dates_with_time = sorted([d for d, m in schedule.items() if m > 0])

        headers = ["実施日", subject, "未割当"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        _write_rows(ws, dates_with_time, [subject],
                    assigned, unassigned, border)


def _write_rows(ws, dates_with_time, subjects, assigned, unassigned, border):
    """行データを書き込む共通関数"""
    order = {"復習": 0, "定着": 1, "再定着": 2, "予習": 3}

    for row_idx, d in enumerate(dates_with_time, start=2):
        d_obj = date.fromisoformat(d)
        date_label = (f"{d_obj.month}/{d_obj.day}"
                      f"（{DOW_JA[d_obj.weekday()]}）")

        date_cell = ws.cell(row=row_idx, column=1, value=date_label)
        date_cell.font = Font(bold=True)
        date_cell.fill = PatternFill("solid", fgColor="E8EEF8")
        date_cell.alignment = Alignment(horizontal="center", vertical="top")
        date_cell.border = border

        day_items = [item for item in assigned if item["assigned_date"] == d]

        for col_idx, subject in enumerate(subjects, start=2):
            subject_items = sorted(
                [item for item in day_items if item["subject"] == subject],
                key=lambda x: order.get(x["category"], 9)
            )
            cell_text = "\\n".join([make_cell_text(i) for i in subject_items])
            color = CATEGORY_COLORS.get(
                subject_items[0]["category"], "FFFFFF") if subject_items else "FFFFFF"

            cell = ws.cell(row=row_idx, column=col_idx, value=cell_text)
            if subject_items:
                cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)
            cell.border = border

        unassigned_cell = ws.cell(
            row=row_idx, column=len(subjects) + 2, value="")
        unassigned_cell.border = border

    if unassigned:
        unassigned_row = len(dates_with_time) + 2
        label_cell = ws.cell(row=unassigned_row, column=1, value="未割当")
        label_cell.font = Font(bold=True)
        label_cell.fill = PatternFill("solid", fgColor="F0F0F0")
        label_cell.alignment = Alignment(horizontal="center", vertical="top")
        label_cell.border = border

        for col_idx, subject in enumerate(subjects, start=2):
            subject_items = [i for i in unassigned if i["subject"] == subject]
            cell_text = "\\n".join([make_cell_text(i) for i in subject_items])
            cell = ws.cell(row=unassigned_row, column=col_idx, value=cell_text)
            cell.fill = PatternFill("solid", fgColor="F0F0F0")
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)
            cell.border = border

    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(subjects) + 3):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)].width = 35
    for row_idx in range(2, len(dates_with_time) + 3):
        ws.row_dimensions[row_idx].height = 80


def export_excel(target_date_str, output_path, student_id=None,
                 start_date=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    conn = get_connection()
    c = conn.cursor()
    if student_id:
        c.execute("SELECT student_id, name, subjects, plan_mode FROM students "
                  "WHERE student_id=?", (student_id,))
    else:
        c.execute("SELECT student_id, name, subjects, plan_mode FROM students "
                  "ORDER BY student_id")
    students = c.fetchall()
    conn.close()

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    start_date_str = start_date if start_date else date.today().isoformat()

    for student in students:
        sid = student["student_id"]
        name = student["name"]
        subjects = [s.strip() for s in student["subjects"].split(",")]
        plan_mode = student["plan_mode"] if student["plan_mode"] else "all"

        if plan_mode == "per_subject":
            write_sheet_per_subject(wb, sid, name, subjects,
                                    start_date_str, target_date_str, border)
        else:
            ws = wb.create_sheet(title=name)
            write_sheet_all(ws, sid, subjects, start_date_str,
                            target_date_str, border, thin)

    wb.save(output_path)
    return output_path
'''

excel_path = os.path.join(base, "excel_export.py")
with open(excel_path, "w", encoding="utf-8") as f:
    f.write(excel_code)
print("✅ excel_export.py を更新しました")

# ─── 3. students.htmlにplan_mode選択を追加 ─────────────
students_html = """<!DOCTYPE html>
<html lang="ja">
<head>
  <meta charset="UTF-8">
  <title>生徒管理</title>
  <style>
    body { font-family: sans-serif; max-width: 900px; margin: 40px auto; background: #f5f7fa; }
    h1 { color: #2c3e50; }
    .form-box { background: white; padding: 24px; border-radius: 8px; margin-bottom: 24px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
    label { display: block; margin-top: 12px; font-weight: bold; }
    input[type="text"] { width: 100%; padding: 8px; margin-top: 4px; border: 1px solid #ccc; border-radius: 4px; font-size: 15px; box-sizing: border-box; }
    .hint { font-size: 12px; color: #888; margin-top: 4px; min-height: 1em; }
    .row3 { display: grid; grid-template-columns: 1fr 2fr 2fr; gap: 12px; align-items: start; }
    .add-btn-row { margin-top: 16px; }
    button { padding: 10px 28px; color: white; border: none; border-radius: 6px; font-size: 15px; cursor: pointer; }
    .btn-add { background: #4472C4; }
    .btn-add:hover { background: #2c5fa8; }
    table { width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
    th { background: #4472C4; color: white; padding: 10px; }
    td { padding: 10px; border-bottom: 1px solid #eee; text-align: center; }
    .edit-btn { padding: 4px 12px; background: #4472C4; color: white; border: none; border-radius: 4px; font-size: 12px; cursor: pointer; margin-right: 4px; }
    .edit-btn:hover { background: #2c5fa8; }
    .del-btn { padding: 4px 12px; background: #e74c3c; color: white; border: none; border-radius: 4px; font-size: 12px; cursor: pointer; }
    .del-btn:hover { background: #c0392b; }
    .edit-form input, .edit-form select { padding: 4px; font-size: 13px; width: 100%; box-sizing: border-box; }
    a { color: #4472C4; }
    .badge-all { background: #4472C4; color: white; padding: 2px 8px; border-radius: 4px; font-size: 11px; }
    .badge-per { background: #27ae60; color: white; padding: 2px 8px; border-radius: 4px; font-size: 11px; }
  </style>
</head>
<body>
  <h1>⑦ 生徒管理</h1>
  <div class="form-box">
    <h2>生徒を追加</h2>
    <form method="POST">
      <input type="hidden" name="action" value="add">
      <div class="row3">
        <div>
          <label>生徒ID</label>
          <input type="text" name="student_id" placeholder="例：S005" required>
          <p class="hint">&nbsp;</p>
        </div>
        <div>
          <label>氏名</label>
          <input type="text" name="name" placeholder="例：山田太郎" required>
          <p class="hint">&nbsp;</p>
        </div>
        <div>
          <label>担当教科</label>
          <input type="text" name="subjects" placeholder="例：数学,英語" required>
          <p class="hint">カンマ区切りで入力</p>
        </div>
      </div>
      <div class="add-btn-row">
        <button type="submit" class="btn-add">追加する</button>
      </div>
    </form>
  </div>

  <table>
    <tr>
      <th>生徒ID</th><th>氏名</th><th>担当教科</th><th>計画モード</th><th>操作</th>
    </tr>
    {% for s in students %}
    <tr>
      <form method="POST" class="edit-form">
        <input type="hidden" name="action" value="edit">
        <input type="hidden" name="student_id" value="{{ s.student_id }}">
        <td>{{ s.student_id }}</td>
        <td><input type="text" name="name" value="{{ s.name }}"></td>
        <td><input type="text" name="subjects" value="{{ s.subjects }}"></td>
        <td>
          <select name="plan_mode" style="padding:4px;font-size:13px;">
            <option value="all"
              {% if s.plan_mode == 'all' or not s.plan_mode %}selected{% endif %}>
              全教科一括
            </option>
            <option value="per_subject"
              {% if s.plan_mode == 'per_subject' %}selected{% endif %}>
              教科ごと個別
            </option>
          </select>
        </td>
        <td>
          <button type="submit" class="edit-btn">更新</button>
      </form>
      <form method="POST" style="display:inline"
            onsubmit="return confirm('{{ s.name }}を削除しますか？関連する全データが削除されます。')">
        <input type="hidden" name="action" value="delete">
        <input type="hidden" name="student_id" value="{{ s.student_id }}">
        <button class="del-btn">削除</button>
      </form>
        </td>
    </tr>
    {% endfor %}
  </table>
  <p><a href="/">← トップに戻る</a></p>
</body>
</html>"""

with open(os.path.join(templates, "students.html"), "w", encoding="utf-8") as f:
    f.write(students_html)
print("✅ students.html を更新しました")

# ─── 4. app.pyのstudents関数にplan_modeを追加 ───────────
app_path = os.path.join(base, "app.py")
with open(app_path, "r", encoding="utf-8") as f:
    app_content = f.read()

old_edit = '''        elif action == "edit":
            student_id = request.form["student_id"]
            name = request.form["name"]
            subjects = request.form["subjects"]
            c.execute("""
                UPDATE students SET name=?, subjects=? WHERE student_id=?
            """, (name, subjects, student_id))'''

new_edit = '''        elif action == "edit":
            student_id = request.form["student_id"]
            name = request.form["name"]
            subjects = request.form["subjects"]
            plan_mode = request.form.get("plan_mode", "all")
            c.execute("""
                UPDATE students SET name=?, subjects=?, plan_mode=?
                WHERE student_id=?
            """, (name, subjects, plan_mode, student_id))'''

if old_edit in app_content:
    app_content = app_content.replace(old_edit, new_edit)
    with open(app_path, "w", encoding="utf-8") as f:
        f.write(app_content)
    print("✅ app.py を更新しました")
else:
    print("❌ app.pyの対象箇所が見つかりません")

print("✅ 全て完了")