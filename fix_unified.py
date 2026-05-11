import os

base = r"C:\Users\ynaka\study_planner"
templates = os.path.join(base, "templates")

# ─── excel_export.py を完全に書き直す ─────────────────
excel_code = '''import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database import (get_connection, get_plan_v2, get_schedule,
                      get_class_dates_in_range)
from datetime import date, timedelta
from collections import defaultdict

DOW_JA = ["月", "火", "水", "木", "金", "土", "日"]

CATEGORY_COLORS = {
    "予習":   "DDEEFF",
    "復習":   "DDFFDD",
    "定着":   "FFFADD",
    "再定着": "FFE5DD",
    "未割当": "F0F0F0",
}

MASTERY_MULTIPLIER = {1: 1.3, 2: 1.0, 3: 0.5}
CATEGORY_GROUP = {"復習": 0, "定着": 0, "再定着": 0, "予習": 1}


def get_adjusted_minutes(item):
    mastery_level = int(item.get("mastery_int", 1) or 1)
    mastery_level = max(1, min(3, mastery_level))
    multiplier = MASTERY_MULTIPLIER.get(mastery_level, 1.0)
    return max(5, round(item["estimated_minutes"] * multiplier / 5) * 5)


def priority_score(item):
    group = CATEGORY_GROUP.get(item["category"], 1)
    mastery = int(item.get("mastery_int", 1) or 1)
    review_value = int(item.get("review_value", 3) or 3)
    importance = int(item.get("importance", 3) or 3)
    difficulty = int(item.get("difficulty", 3) or 3)
    return (group, mastery, -review_value, -importance, difficulty)


def assign_days_v2(plan, schedule, student_id, start_date_str, end_date_str):
    """優先度×テキスト分散を考慮した割り振りロジック"""
    remaining = {d: m for d, m in schedule.items() if m > 0}
    dates_sorted = sorted(remaining.keys())

    if not dates_sorted:
        for item in plan:
            item["assigned_date"] = None
        return [], plan

    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        SELECT DISTINCT p.subject FROM problems p
        JOIN assignments a ON p.problem_id = a.problem_id
        WHERE a.student_id=?
    """, (student_id,))
    subjects = [r["subject"] for r in c.fetchall()]
    conn.close()

    subject_class_dates = {
        s: get_class_dates_in_range(student_id, s, start_date_str, end_date_str)
        for s in subjects
    }

    assigned = []
    unassigned = []
    date_textbook_count = defaultdict(lambda: defaultdict(int))

    def try_assign(item, search_dates):
        minutes = get_adjusted_minutes(item)
        textbook = item["textbook"]
        valid_dates = [d for d in search_dates if remaining.get(d, 0) >= minutes]
        sorted_dates = sorted(valid_dates,
                              key=lambda d: date_textbook_count[d][textbook])
        for d in sorted_dates:
            remaining[d] -= minutes
            item["assigned_date"] = d
            date_textbook_count[d][textbook] += 1
            assigned.append(item)
            return True
        return False

    plan_sorted = sorted(plan, key=priority_score)
    review_teichaku = [p for p in plan_sorted
                       if p["category"] in ("復習", "定着", "再定着")]
    yosyu = [p for p in plan_sorted if p["category"] == "予習"]

    for item in [p for p in review_teichaku if p["category"] == "復習"]:
        class_dates = subject_class_dates.get(item["subject"], [])
        search_from = dates_sorted[0]
        past_classes = [d for d in class_dates if d <= start_date_str]
        if past_classes:
            search_from = past_classes[-1]
        search_dates = [d for d in dates_sorted if d >= search_from]
        if not try_assign(item, search_dates):
            item["assigned_date"] = None
            unassigned.append(item)

    for item in [p for p in review_teichaku
                 if p["category"] in ("定着", "再定着")]:
        if not try_assign(item, dates_sorted):
            item["assigned_date"] = None
            unassigned.append(item)

    for item in yosyu:
        class_dates = subject_class_dates.get(item["subject"], [])
        future_classes = sorted([d for d in class_dates if d > start_date_str])
        if future_classes:
            next_class = future_classes[0]
            search_dates = list(reversed(
                [d for d in dates_sorted if d <= next_class]))
        else:
            search_dates = list(reversed(dates_sorted))
        if not try_assign(item, search_dates):
            item["assigned_date"] = None
            unassigned.append(item)

    return assigned, unassigned


def build_plan_data(student_id, start_date_str, target_date_str,
                    subject_filter=None):
    """
    プレビュー・Excel共通の計画データ生成関数
    subject_filter: 特定教科のみ出力する場合に指定
    """
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT student_id, name, subjects, plan_mode FROM students "
              "WHERE student_id=?", (student_id,))
    student = c.fetchone()
    conn.close()

    if not student:
        return None

    all_subjects = [s.strip() for s in student["subjects"].split(",")]
    plan_mode = student["plan_mode"] if student["plan_mode"] else "all"

    schedule = get_schedule(student_id, start_date_str, target_date_str)
    all_plan = get_plan_v2(student_id, start_date_str, target_date_str)

    if subject_filter:
        subjects = [subject_filter]
        plan = [p for p in all_plan if p["subject"] == subject_filter]
    elif plan_mode == "per_subject":
        subjects = all_subjects
        plan = all_plan
    else:
        subjects = all_subjects
        plan = all_plan

    assigned, unassigned = assign_days_v2(
        plan, schedule, student_id, start_date_str, target_date_str)

    dates_with_time = sorted([d for d, m in schedule.items() if m > 0])

    order = {"復習": 0, "定着": 1, "再定着": 2, "予習": 3}

    rows = []
    for d in dates_with_time:
        d_obj = date.fromisoformat(d)
        date_label = (f"{d_obj.month}/{d_obj.day}"
                      f"（{DOW_JA[d_obj.weekday()]}）")
        day_items = [item for item in assigned if item["assigned_date"] == d]
        row = {"date": date_label, "date_str": d, "subjects": {}}
        for subject in subjects:
            subject_items = sorted(
                [item for item in day_items if item["subject"] == subject],
                key=lambda x: order.get(x["category"], 9)
            )
            row["subjects"][subject] = subject_items
        rows.append(row)

    unassigned_by_subject = {}
    for subject in subjects:
        unassigned_by_subject[subject] = [
            item for item in unassigned if item["subject"] == subject]

    return {
        "student_name": student["name"],
        "student_id": student_id,
        "subjects": subjects,
        "plan_mode": plan_mode,
        "rows": rows,
        "unassigned": unassigned_by_subject,
        "schedule": schedule,
    }


def export_excel(target_date_str, output_path, student_id=None,
                 start_date=None, subject_filter=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    conn = get_connection()
    c = conn.cursor()
    if student_id:
        c.execute("SELECT student_id, name, subjects, plan_mode FROM students "
                  "WHERE student_id=?", (student_id,))
    else:
        c.execute("SELECT student_id, name, subjects, plan_mode FROM students "
                  "ORDER BY student_id")
    students = c.fetchall()
    conn.close()

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    start_date_str = start_date if start_date else date.today().isoformat()

    for student in students:
        sid = student["student_id"]
        name = student["name"]
        plan_mode = student["plan_mode"] if student["plan_mode"] else "all"

        if subject_filter:
            subjects_to_process = [subject_filter]
        elif plan_mode == "per_subject":
            subjects_to_process = [s.strip()
                                   for s in student["subjects"].split(",")]
        else:
            subjects_to_process = None  # 全教科一括

        if subjects_to_process:
            # 教科ごとに別シートで出力
            for subject in subjects_to_process:
                data = build_plan_data(
                    sid, start_date_str, target_date_str, subject)
                if not data:
                    continue
                ws = wb.create_sheet(title=f"{name}_{subject}")
                _write_excel_sheet(ws, data, border)
        else:
            # 全教科一括で1シートに出力
            data = build_plan_data(sid, start_date_str, target_date_str)
            if not data:
                continue
            ws = wb.create_sheet(title=name)
            _write_excel_sheet(ws, data, border)

    wb.save(output_path)
    return output_path


def _write_excel_sheet(ws, data, border):
    """Excelシートに計画データを書き込む"""
    subjects = data["subjects"]
    rows = data["rows"]
    unassigned = data["unassigned"]

    # ヘッダー：実施日 | 教科ごとの問題番号 | 教科ごとの学習指示 | 未割当
    header_cols = ["実施日"]
    for subject in subjects:
        header_cols.append(f"{subject}（問題）")
        header_cols.append(f"{subject}（指示）")
    header_cols.append("未割当")

    for col, h in enumerate(header_cols, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    order = {"復習": 0, "定着": 1, "再定着": 2, "予習": 3}

    for row_idx, row in enumerate(rows, start=2):
        # 実施日セル
        date_cell = ws.cell(row=row_idx, column=1, value=row["date"])
        date_cell.font = Font(bold=True)
        date_cell.fill = PatternFill("solid", fgColor="E8EEF8")
        date_cell.alignment = Alignment(horizontal="center", vertical="top")
        date_cell.border = border

        col_idx = 2
        for subject in subjects:
            subject_items = sorted(
                row["subjects"].get(subject, []),
                key=lambda x: order.get(x["category"], 9)
            )

            # 問題番号列
            problem_lines = []
            instruction_lines = []
            for item in subject_items:
                problem_lines.append(
                    f"【{item['category']}】{item['textbook']} "
                    f"{item['problem_number']}")
                instruction_lines.append(item.get("instruction", "") or "")

            problem_text = "\n".join(problem_lines)
            instruction_text = "\n".join(instruction_lines)

            color = "FFFFFF"
            if subject_items:
                color = CATEGORY_COLORS.get(
                    subject_items[0]["category"], "FFFFFF")

            prob_cell = ws.cell(row=row_idx, column=col_idx,
                                value=problem_text)
            if subject_items:
                prob_cell.fill = PatternFill("solid", fgColor=color)
            prob_cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)
            prob_cell.border = border
            col_idx += 1

            instr_cell = ws.cell(row=row_idx, column=col_idx,
                                 value=instruction_text)
            if subject_items:
                instr_cell.fill = PatternFill("solid", fgColor=color)
            instr_cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)
            instr_cell.border = border
            col_idx += 1

        # 未割当列（空）
        unassigned_cell = ws.cell(row=row_idx, column=col_idx, value="")
        unassigned_cell.border = border

    # 未割当行
    all_unassigned = []
    for subject in subjects:
        all_unassigned.extend(unassigned.get(subject, []))

    if all_unassigned:
        unassigned_row = len(rows) + 2
        label_cell = ws.cell(row=unassigned_row, column=1, value="未割当")
        label_cell.font = Font(bold=True)
        label_cell.fill = PatternFill("solid", fgColor="F0F0F0")
        label_cell.alignment = Alignment(horizontal="center", vertical="top")
        label_cell.border = border

        col_idx = 2
        for subject in subjects:
            subject_items = unassigned.get(subject, [])
            problem_text = "\n".join(
                [f"【{i['category']}】{i['textbook']} {i['problem_number']}"
                 for i in subject_items])
            instruction_text = "\n".join(
                [i.get("instruction", "") or "" for i in subject_items])

            prob_cell = ws.cell(row=unassigned_row, column=col_idx,
                                value=problem_text)
            prob_cell.fill = PatternFill("solid", fgColor="F0F0F0")
            prob_cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)
            prob_cell.border = border
            col_idx += 1

            instr_cell = ws.cell(row=unassigned_row, column=col_idx,
                                 value=instruction_text)
            instr_cell.fill = PatternFill("solid", fgColor="F0F0F0")
            instr_cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)
            instr_cell.border = border
            col_idx += 1

    # 列幅設定
    ws.column_dimensions["A"].width = 14
    col_letter_idx = 2
    for subject in subjects:
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_letter_idx)].width = 30
        col_letter_idx += 1
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_letter_idx)].width = 35
        col_letter_idx += 1

    for row_idx in range(2, len(rows) + 3):
        ws.row_dimensions[row_idx].height = 80
'''

excel_path = os.path.join(base, "excel_export.py")
with open(excel_path, "w", encoding="utf-8") as f:
    f.write(excel_code)
print("✅ excel_export.py を完全に書き直しました")

# ─── app.pyのpreview・exportを修正 ─────────────────────
app_path = os.path.join(base, "app.py")
with open(app_path, "r", encoding="utf-8") as f:
    app_content = f.read()

# preview関数を完全に置き換え
old_preview = '''@app.route("/preview", methods=["GET", "POST"])
def preview():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM students ORDER BY student_id")
    students = c.fetchall()
    conn.close()

    preview_data = None
    selected = {}

    if request.method == "POST":
        student_id = request.form["student_id"]
        target_date = request.form["date"]
        start_date = request.form.get("start_date", "")
        if not start_date:
            start_date = date.today().isoformat()

        selected = {
            "student_id": student_id,
            "date": target_date,
            "start_date": start_date,
        }

        from database import get_plan, get_schedule
        from excel_export import assign_days_v2 as assign_days

        conn2 = get_connection()
        c2 = conn2.cursor()
        c2.execute("SELECT name, subjects FROM students WHERE student_id=?",
                   (student_id,))
        student = c2.fetchone()
        conn2.close()

        subjects = [s.strip() for s in student["subjects"].split(",")]
        schedule = get_schedule(student_id, start_date, target_date)
        plan = get_plan(student_id, target_date)
        assigned, unassigned = assign_days(plan, schedule, student_id, start_date, target_date)

        dates_with_time = sorted([d for d, m in schedule.items() if m > 0])

        DOW_JA = ["月", "火", "水", "木", "金", "土", "日"]
        rows = []
        for d in dates_with_time:
            d_obj = date.fromisoformat(d)
            date_label = f"{d_obj.month}/{d_obj.day}（{DOW_JA[d_obj.weekday()]}）"
            day_items = [item for item in assigned if item["assigned_date"] == d]
            row = {"date": date_label, "subjects": {}}
            for subject in subjects:
                subject_items = [item for item in day_items if item["subject"] == subject]
                row["subjects"][subject] = subject_items
            rows.append(row)

        unassigned_by_subject = {}
        for subject in subjects:
            unassigned_by_subject[subject] = [
                item for item in unassigned if item["subject"] == subject]

        preview_data = {
            "student_name": student["name"],
            "subjects": subjects,
            "rows": rows,
            "unassigned": unassigned_by_subject,
        }

    return render_template("preview.html", students=students,
                           preview_data=preview_data, selected=selected)'''

new_preview = '''@app.route("/preview", methods=["GET", "POST"])
def preview():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM students ORDER BY student_id")
    students = c.fetchall()
    conn.close()

    preview_data = None
    selected = {}

    if request.method == "POST":
        student_id = request.form["student_id"]
        target_date = request.form["date"]
        start_date = request.form.get("start_date", "")
        subject_filter = request.form.get("subject_filter", "")
        if not start_date:
            start_date = date.today().isoformat()

        selected = {
            "student_id": student_id,
            "date": target_date,
            "start_date": start_date,
            "subject_filter": subject_filter,
        }

        from excel_export import build_plan_data
        preview_data = build_plan_data(
            student_id, start_date, target_date,
            subject_filter if subject_filter else None)

    # 各生徒の教科リストを渡す
    conn2 = get_connection()
    c2 = conn2.cursor()
    c2.execute("SELECT student_id, subjects, plan_mode FROM students")
    student_subjects = {
        r["student_id"]: {
            "subjects": [s.strip() for s in r["subjects"].split(",")],
            "plan_mode": r["plan_mode"] or "all"
        }
        for r in c2.fetchall()
    }
    conn2.close()

    return render_template("preview.html", students=students,
                           preview_data=preview_data, selected=selected,
                           student_subjects=student_subjects)'''

if old_preview in app_content:
    app_content = app_content.replace(old_preview, new_preview)
    print("✅ preview関数を更新しました")
else:
    print("❌ preview関数が見つかりません。手動で確認が必要です")

# export関数にsubject_filterを追加
old_export_call = '''        export_excel(target_date, output_path,
                     student_id=student_id, start_date=start_date)'''
new_export_call = '''        subject_filter = request.form.get("subject_filter", "")
        export_excel(target_date, output_path,
                     student_id=student_id, start_date=start_date,
                     subject_filter=subject_filter if subject_filter else None)'''

if old_export_call in app_content:
    app_content = app_content.replace(old_export_call, new_export_call)
    print("✅ export関数を更新しました")

# export_pdf関数にsubject_filterを追加
old_pdf_call = '''    export_pdf(target_date, output_path,
               student_id=student_id, start_date=start_date)'''
new_pdf_call = '''    subject_filter = request.form.get("subject_filter", "")
    export_pdf(target_date, output_path,
               student_id=student_id, start_date=start_date,
               subject_filter=subject_filter if subject_filter else None)'''

if old_pdf_call in app_content:
    app_content = app_content.replace(old_pdf_call, new_pdf_call)
    print("✅ export_pdf関数を更新しました")

# 過去の計画表削除＋履歴リセット機能を追加
history_reset_route = '''
@app.route("/history/delete/<int:history_id>", methods=["POST"])
def history_delete(history_id):
    reset_assignments = request.form.get("reset_assignments") == "1"
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM plan_history WHERE history_id=?", (history_id,))
    row = c.fetchone()
    if row and reset_assignments:
        student_id = row["student_id"]
        start_date = row["start_date"]
        end_date = row["end_date"]
        # 当該期間の出題予定をリセット
        c.execute("""
            DELETE FROM assignments
            WHERE student_id=?
            AND scheduled_date BETWEEN ? AND ?
        """, (student_id, start_date, end_date))
        # 当該期間の授業記録をリセット
        c.execute("""
            DELETE FROM history
            WHERE student_id=?
            AND date BETWEEN ? AND ?
        """, (student_id, start_date, end_date))
        print(f"リセット完了：{student_id} {start_date}〜{end_date}")
    # Excelファイルを削除
    if row:
        import os as _os
        if _os.path.exists(row["excel_path"]):
            _os.remove(row["excel_path"])
    c.execute("DELETE FROM plan_history WHERE history_id=?", (history_id,))
    conn.commit()
    conn.close()
    return redirect("/history")

'''

if "/history/delete/" not in app_content:
    idx = app_content.rfind("if __name__")
    app_content = app_content[:idx] + history_reset_route + app_content[idx:]
    print("✅ 履歴削除ルートを追加しました")

with open(app_path, "w", encoding="utf-8") as f:
    f.write(app_content)

# ─── preview.htmlを更新（教科選択追加） ─────────────────
preview_html = """<!DOCTYPE html>
<html lang="ja">
<head>
  <meta charset="UTF-8">
  <title>計画表プレビュー</title>
  <style>
    body { font-family: sans-serif; max-width: 1200px; margin: 40px auto; background: #f5f7fa; }
    h1 { color: #2c3e50; }
    .form-box { background: white; padding: 24px; border-radius: 8px; margin-bottom: 24px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
    label { display: block; margin-top: 12px; font-weight: bold; }
    select, input[type="date"] { width: 100%; padding: 8px; margin-top: 4px; border: 1px solid #ccc; border-radius: 4px; font-size: 15px; box-sizing: border-box; }
    .btn-row { display: flex; gap: 12px; margin-top: 16px; flex-wrap: wrap; }
    button { padding: 10px 24px; color: white; border: none; border-radius: 6px; font-size: 15px; cursor: pointer; }
    .btn-preview { background: #4472C4; }
    .btn-preview:hover { background: #2c5fa8; }
    .btn-excel { background: #27ae60; }
    .btn-excel:hover { background: #1e8449; }
    .btn-pdf { background: #e74c3c; }
    .btn-pdf:hover { background: #c0392b; }
    .preview-box { background: white; border-radius: 8px; padding: 24px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); overflow-x: auto; }
    h2 { color: #2c3e50; margin-bottom: 16px; }
    table { border-collapse: collapse; width: 100%; }
    th { background: #4472C4; color: white; padding: 10px 14px; text-align: center; white-space: nowrap; }
    td { border: 1px solid #ddd; padding: 8px 10px; vertical-align: top; min-width: 160px; }
    td.date-cell { background: #e8eef8; font-weight: bold; text-align: center; white-space: nowrap; width: 100px; }
    td.unassigned-label { background: #f0f0f0; font-weight: bold; text-align: center; }
    .problem-item { margin-bottom: 8px; padding: 6px 8px; border-radius: 4px; font-size: 13px; line-height: 1.5; }
    .cat-予習  { background: #DDEEFF; }
    .cat-復習  { background: #DDFFDD; }
    .cat-定着  { background: #FFFADD; }
    .cat-再定着 { background: #FFE5DD; }
    .cat-label { font-weight: bold; }
    .instruction { color: #555; font-size: 12px; margin-top: 2px; }
    p.note { color: #888; font-size: 13px; margin-top: 6px; }
    a { color: #4472C4; }
    #subject-filter-row { display: none; }
  </style>
</head>
<body>
  <h1>⑤ 計画表プレビュー・出力</h1>
  <div class="form-box">
    <form method="POST" id="main-form" action="/preview">
      <label>生徒</label>
      <select name="student_id" id="student-select" onchange="updateSubjectFilter()">
        {% for s in students %}
        <option value="{{ s.student_id }}"
          {% if selected.get('student_id') == s.student_id %}selected{% endif %}>
          {{ s.name }}
        </option>
        {% endfor %}
      </select>

      <div id="subject-filter-row">
        <label>教科を選択（教科ごと個別モードの場合）</label>
        <select name="subject_filter" id="subject-filter">
          <option value="">全教科</option>
        </select>
      </div>

      <label>次回授業日（計画終了日）</label>
      <input type="date" name="date" value="{{ selected.get('date', '') }}" required>

      <label>計画開始日（任意・空欄で今日）</label>
      <input type="date" name="start_date" value="{{ selected.get('start_date', '') }}">
      <p class="note">空欄の場合は今日の日付が自動的に使われます。</p>

      <div class="btn-row">
        <button type="submit" class="btn-preview"
          onclick="document.getElementById('main-form').action='/preview';">
          🔍 プレビュー
        </button>
        <button type="submit" class="btn-excel"
          onclick="document.getElementById('main-form').action='/export';">
          📊 Excelダウンロード
        </button>
        <button type="submit" class="btn-pdf"
          onclick="document.getElementById('main-form').action='/export_pdf';">
          📄 PDFダウンロード
        </button>
      </div>
    </form>
  </div>

  {% if preview_data %}
  <div class="preview-box">
    <h2>{{ preview_data.student_name }} の学習計画</h2>
    <table>
      <tr>
        <th>実施日</th>
        {% for subject in preview_data.subjects %}
        <th>{{ subject }}（問題）</th>
        <th>{{ subject }}（指示）</th>
        {% endfor %}
      </tr>
      {% for row in preview_data.rows %}
      <tr>
        <td class="date-cell">{{ row.date }}</td>
        {% for subject in preview_data.subjects %}
        <td>
          {% for item in row.subjects.get(subject, []) %}
          <div class="problem-item cat-{{ item.category }}">
            <div class="cat-label">【{{ item.category }}】{{ item.textbook }} {{ item.problem_number }}</div>
          </div>
          {% endfor %}
        </td>
        <td>
          {% for item in row.subjects.get(subject, []) %}
          <div class="problem-item cat-{{ item.category }}">
            <div class="instruction">{{ item.instruction or "―" }}</div>
          </div>
          {% endfor %}
        </td>
        {% endfor %}
      </tr>
      {% endfor %}
      {% set has_unassigned = [] %}
      {% for subject in preview_data.subjects %}
        {% if preview_data.unassigned.get(subject) %}
          {% set _ = has_unassigned.append(1) %}
        {% endif %}
      {% endfor %}
      {% if has_unassigned %}
      <tr>
        <td class="unassigned-label">未割当</td>
        {% for subject in preview_data.subjects %}
        <td>
          {% for item in preview_data.unassigned.get(subject, []) %}
          <div class="problem-item cat-{{ item.category }}">
            <div class="cat-label">【{{ item.category }}】{{ item.textbook }} {{ item.problem_number }}</div>
          </div>
          {% endfor %}
        </td>
        <td>
          {% for item in preview_data.unassigned.get(subject, []) %}
          <div class="problem-item cat-{{ item.category }}">
            <div class="instruction">{{ item.instruction or "―" }}</div>
          </div>
          {% endfor %}
        </td>
        {% endfor %}
      </tr>
      {% endif %}
    </table>
  </div>
  {% endif %}

  <p><a href="/">← トップに戻る</a></p>

  <script>
    const studentSubjects = {{ student_subjects | tojson }};

    function updateSubjectFilter() {
      const studentId = document.getElementById('student-select').value;
      const info = studentSubjects[studentId];
      const filterRow = document.getElementById('subject-filter-row');
      const filterSelect = document.getElementById('subject-filter');

      if (info && info.plan_mode === 'per_subject') {
        filterRow.style.display = 'block';
        filterSelect.innerHTML = '<option value="">全教科</option>';
        info.subjects.forEach(s => {
          const opt = document.createElement('option');
          opt.value = s;
          opt.textContent = s;
          filterSelect.appendChild(opt);
        });
      } else {
        filterRow.style.display = 'none';
        filterSelect.value = '';
      }
    }

    // 初期表示
    updateSubjectFilter();

    // 選択済みの教科を復元
    const selectedSubject = "{{ selected.get('subject_filter', '') }}";
    if (selectedSubject) {
      document.getElementById('subject-filter').value = selectedSubject;
    }
  </script>
</body>
</html>"""

with open(os.path.join(templates, "preview.html"), "w", encoding="utf-8") as f:
    f.write(preview_html)
print("✅ preview.html を更新しました")

# ─── history.htmlに削除ボタンを追加 ────────────────────
history_html = """<!DOCTYPE html>
<html lang="ja">
<head>
  <meta charset="UTF-8">
  <title>過去の計画表</title>
  <style>
    body { font-family: sans-serif; max-width: 900px; margin: 40px auto; background: #f5f7fa; }
    h1 { color: #2c3e50; }
    .filter-box { background: white; padding: 16px 24px; border-radius: 8px; margin-bottom: 24px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
    select { padding: 8px; border: 1px solid #ccc; border-radius: 4px; font-size: 15px; }
    table { width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
    th { background: #4472C4; color: white; padding: 10px; }
    td { padding: 10px; border-bottom: 1px solid #eee; text-align: center; }
    .dl-btn { padding: 6px 16px; background: #27ae60; color: white; border-radius: 4px; text-decoration: none; font-size: 13px; margin-right: 4px; }
    .dl-btn:hover { background: #1e8449; }
    .del-btn { padding: 6px 12px; background: #e74c3c; color: white; border: none; border-radius: 4px; font-size: 12px; cursor: pointer; margin-right: 2px; }
    .del-btn:hover { background: #c0392b; }
    .reset-btn { padding: 6px 12px; background: #8e44ad; color: white; border: none; border-radius: 4px; font-size: 12px; cursor: pointer; }
    .reset-btn:hover { background: #6c3483; }
    a.back { color: #4472C4; }
  </style>
</head>
<body>
  <h1>⑧ 過去の計画表</h1>
  <div class="filter-box">
    <form method="GET">
      <label>生徒で絞り込み：</label>
      <select name="student_id" onchange="this.form.submit()">
        <option value="">全生徒</option>
        {% for s in students %}
        <option value="{{ s.student_id }}"
          {% if s.student_id == selected_student %}selected{% endif %}>
          {{ s.name }}
        </option>
        {% endfor %}
      </select>
    </form>
  </div>
  <table>
    <tr>
      <th>生徒</th><th>生成日</th><th>計画開始日</th><th>計画終了日</th><th>操作</th>
    </tr>
    {% for h in histories %}
    <tr>
      <td>{{ h.name }}</td>
      <td>{{ h.generated_date }}</td>
      <td>{{ h.start_date }}</td>
      <td>{{ h.end_date }}</td>
      <td>
        <a class="dl-btn" href="/history/download/{{ h.history_id }}">📥 DL</a>
        <form method="POST" action="/history/delete/{{ h.history_id }}"
              style="display:inline"
              onsubmit="return confirm('この計画表を削除しますか？')">
          <input type="hidden" name="reset_assignments" value="0">
          <button class="del-btn">削除</button>
        </form>
        <form method="POST" action="/history/delete/{{ h.history_id }}"
              style="display:inline"
              onsubmit="return confirm('計画表と出題履歴をリセットしますか？この操作は取り消せません。')">
          <input type="hidden" name="reset_assignments" value="1">
          <button class="reset-btn">🔄 履歴リセット</button>
        </form>
      </td>
    </tr>
    {% else %}
    <tr><td colspan="5">計画表の履歴がありません</td></tr>
    {% endfor %}
  </table>
  <p><a class="back" href="/">← トップに戻る</a></p>
</body>
</html>"""

with open(os.path.join(templates, "history.html"), "w", encoding="utf-8") as f:
    f.write(history_html)
print("✅ history.html を更新しました")
print("✅ 全て完了")